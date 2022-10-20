
contador = {   "Clientes": 0,
               "Salas": 0,
               "Programacion": 30
           }

cliente =  {   "CliCod": 0,
               "CliNam": 'Por Definir'              
           }


sala =     {   "sala 0": 'Sala Cero'       
           }

salaCapacidad =     {   "sala 0": '0'       
           }


agenda =   {   "0": '0'       
           }

import openpyxl
import csv

#### Validacion que no se repita Fecha-Sala-Turno
def f_EncontroCita(cita,agenda):
                                                    
                  v_validacion = False

                  for clave in agenda:
                      if (len(clave) >= 10):
                          if clave[0:10] == cita:
                             v_validacion = True

                  return v_validacion


def f_ValidaCadenaVacia(cadena):
  if (len(cadena) == 0):
    return True
  else:
    return False

def p_EscribeCsv(v_Archivo,v_NombreColumnas,v_Diccionario):
  with open(v_Archivo, 'w') as csvfile:
      writer = csv.DictWriter(csvfile, fieldnames = v_NombreColumnas)
      writer.writeheader()
      writer.writerows(v_Diccionario)  


#####Escritura de Archivos
def p_EscribirARchivos():
  v_Columnas = ['Clientes','Salas','Programacion']
  p_EscribeCsv('Contador.csv',v_Columnas,contador)

  v_Columnas = ['CliCod','CliNam']
  p_EscribeCsv('Clientes.csv',v_Columnas,cliente)


  v_Columnas = ['SalCod','SalNam']
  p_EscribeCsv('Salas.csv',v_Columnas,sala)

  v_Columnas = ['SalCod','SalCap']
  p_EscribeCsv('SalaCap.csv',v_Columnas,salaCapacidad)


  v_Columnas = ['Agecod','AgeEvento']
  p_EscribeCsv('Agenda.csv',v_Columnas,agenda)

#####


opcion = '0'
print("") #Espacio
while not(opcion=='9'):
    print("      **** MENU PRINCIPAL ****  ")
    print(' [1]. Reservaciones ')
    print(' [2]. Reportes')
    print(' [3]. Registrar una sala')
    print(' [4]. Registrar un cliente')
    print(' [9]. Salir')
    print("")

    opcion=input('  --- ¿Cuál opcion?: ')
    
    if (opcion=='1'):
        opcion = '0'

        while not(opcion=='9'):
            print('          *** SUBMENU DE RESERVACIONES***             ')
            print(' [1]. Registrar una nueva reservación ')                     
            print(' [2]. Modificiar descripción de una reservación ') 
            print(' [3]. Consultar disponibilidad de salas para una fecha ')
            print(' [9]. Salir')                                                
            print("") #espacio

            opcion=input('  --- ¿Cuál opcion?: ')
        
######### Sub menu de RESERVACIONES #########        
            if (opcion=='1'):
                print('Registrar una nueva reservación')

                #Validacion inicio SALA
                while True:
                  try:
                    iSala = int(input("SALA ?: "))
                  except ValueError:
                    print("") #Espacio
                    print("Debes escribir un número.")
                    continue

                  if (iSala < 0):
                    print("Debes escribir un número positivo.")
                    continue
                  else:
                    break
                    if (iSala in sala):
                      validar = sala.get(iSala)
                      print(validar)
                      break
                    else:
                      print("No Existe sala " , iSala)
                      continue
                #Validacion fin SALA

                #Validacion inicio CLIENTE
                while True:
                  try:
                    iCliente = int(input("CLIENTE ?: "))
                  except ValueError:
                    print("Debes escribir un número.")
                    continue

                  if (iCliente < 0):
                    print("Debes escribir un número positivo.")
                    continue
                  else:

                    if (iCliente in cliente):
                      validar = cliente.get(iCliente)
                      print(validar)
                      print("") #Espacio
                      break
                    else:
                      print("No Existe cliente " , iCliente)
                      continue
                #Validacion fin CLIENTE

                #Validacion inicio TURNO
                while True:
                    iTurno = input("TURNO [M]añana [T]arde  [N]oche ?: ")

                    iTurno = iTurno.upper()

                    if (iTurno != "M" and iTurno != "T" and iTurno != "N"):
                      print("Turno no valido debe ser [M] [T] [N]")
                      continue
                    else:
                      break
                #Validacion fin TURNO
                      
                #Validacion inicio FECHA
                while True:
                    from datetime import datetime,date,timedelta

                    test_str = input("FECHA AGENDA DD-MM-YY ?: ")
 
                    # initializing format
                    format = "%d-%m-%y"
                    EndDate = datetime.now()+timedelta(days=2)

                    try:
                        res = datetime.strptime(test_str, format)
                        v_abuscar = test_str + str(iSala) + iTurno 

                        if (res >= EndDate) and not f_EncontroCita(v_abuscar,agenda):
                          break
                        else:
                              if not (res >= EndDate):
                               print("Fecha debe ser con 2 dias de Anticipacion")
                               continue
                              else:
                                print("La cita ya existe!!!")
                                continue
                    except ValueError:

                      print("Fecha Invalida ",test_str )
                      continue

                #Validacion FECHA

                #8 Car fecha + 1 Car Sala +  1 Car Turno + 1 Car cliente
                AgendaFolio = test_str + str(iSala) + iTurno + str(iCliente) 
                print("Folio de agenda = ",AgendaFolio)

                AgendaEventoNombre = input("Nombre Evento ?: ")
                agenda[AgendaFolio] = AgendaEventoNombre
                print("")#Espacio
                #print(agenda)


            elif (opcion=='4'):
                print(' **** Consulta por Sala ****')
                iSala=input('¿Numero de Sala?: ')

                for clave in agenda:
                    if (len(clave) >= 10):
                      if (clave[8:10] == str(iSala)):
                        print(clave,"      ",agenda[clave])


                print("") #espacio
            elif (opcion=='2'):
                print('Modificiar descripción de una reservación')
                print("") #espacio

                v_folio = input("Capture el folio de la Agenda a Modificar")
                v_loencontre = False
                v_evento = ""
                v_eventomodificado = ""

                for clave in agenda:
                  if (len(clave) >= 10):
                   if (clave == str(v_folio)):

                      v_loencontre = True
                      v_evento = agenda[clave]

                if v_loencontre:

                  v_eventomodificado = input(" Introduce la nueva descripcion ")
                  agenda[clave] = v_eventomodificado
                  print("La modificacion se realizo correctamente")
                else:
                  print("No existe el folio ", v_folio)


            elif (opcion=='3'):
                print('Consultar disponibilidad de salas para una fecha')
                print("") #espacio
                v_fecha  = input("Dame la fecha ?")

                ##Recorer la Sala
                for clave_sala in sala:
                  v_encontre_M = False
                  v_encontre_T = False
                  v_encontre_N = False

                  for clave_agenda in agenda:
                    if (len(clave_agenda) >= 10):

                       if (clave_agenda[0:8]) == v_fecha and (clave_agenda[8:9]) == str(clave_sala) and (clave_agenda[9:10] == "M"  ):                 
                          v_encontre_M = True

                       if (clave_agenda[0:8]) == v_fecha and (clave_agenda[8:9]) == str(clave_sala) and (clave_agenda[9:10] == "T"  ):                 
                          v_encontre_T = True

                       if (clave_agenda[0:8]) == v_fecha and (clave_agenda[8:9]) == str(clave_sala) and (clave_agenda[9:10] == "N"  ):                 
                          v_encontre_N = True

                  if not (v_encontre_M):
                     print("Para fecha = ",v_fecha,"  y sala   = ",clave_sala , " Esta disponible Turno M")
                      
                  if not (v_encontre_T):
                     print("Para fecha = ",v_fecha,"  y sala   = ",clave_sala , " Esta disponible Turno T")

                  if not (v_encontre_N):
                     print("Para fecha = ",v_fecha,"  y sala   = ",clave_sala , " Esta disponible Turno N")
                     print()

            elif (opcion=='9'):
                print(' ** Saliendo del submenu de reservaciones **')
                print("")#espacio
                opcion = '0'
                break
            else:
                print('No existe la opcion.')

        print("") #Espacio


    if (opcion=='2'):
        opcion = '0' 

######### Sub menu de REPORTES #########
        while not(opcion=='9'):
            print('          *** SUBMENU DE REPORTES***             ')
            print(' [1]. Reporte en pantalla de reservaciones para una fecha')                 
            print(' [2]. Exportar reporte tabular en Excel') 
            print(' [9]. Salir')                                                
            print("") #espacio

            opcion=input('  --- ¿Cuál opcion?: ')
        
            if (opcion=='1'):
                print('[1].Reporte en pantalla de reservaciones para una fecha')
                print("") #espacio

                v_fecha=input('¿Que Fecha ?: ')

                print('*****************************************************')
                print(' **** Reporte de reservaciones el dia ',v_fecha,'****')
                print('Folio            Turno       Nombre Evento')
                print('*****************************************************')
                for clave in agenda:
                    if (len(clave) >= 10):
                      if (clave[0:8] == str(v_fecha)):
                        print(clave,"      ",clave[9:10],"      ",agenda[clave])
                print(' ****************** Fin del reporte *****************')
                print("") #Espacio

            elif (opcion=='2'):
                print('[2].Exportar reporte tabular en Excel')
                print("") #espacio

                libro = openpyxl.Workbook()
                hoja = libro["Sheet"] 

                hoja.title = "Reporte tabular"
                hoja['A1'].value='Reporte de reservaciones el dia '
                hoja['A2'].value=('Sala')
                hoja['B2'].value='Cliente'
                hoja['C2'].value='Folio'
                hoja['D2'].value='Evento'
                hoja['E2'].value='Turno'

                v_renglon = 3
                for clave in agenda:
                    if (len(clave) >= 10):

                          cellref=hoja.cell(v_renglon, column=1)
                          cellref.value= clave[8:9] #NombreSala

                          cellref=hoja.cell(v_renglon, column=2)
                          cellref.value=clave[10:11] #NombreCliente

                          cellref=hoja.cell(v_renglon, column=3)
                          cellref.value=clave
                          #hoja['D2'].value='Evento'

                          cellref=hoja.cell(v_renglon, column=4)
                          cellref.value= agenda[clave] #AgendaEventoNombre

                          cellref=hoja.cell(v_renglon, column=5)
                          cellref.value=clave[9:10]

                          v_renglon = v_renglon + 1

                libro.save('ReporteTabular.xlsx')
                print('Libro creado exitosamente!')

            elif (opcion=='9'):
                print(' ** Saliendo del submenu de reportes  **')
                opcion = '0'
                break
            else:
                print('No existe la opcion..')


    elif (opcion=='4'):
        print(' **** Cliente ****')
 
        print("") #Espacio
        siguiente = contador.get('Clientes')

        # update value
        siguiente =  siguiente + 1
        contador['Clientes'] = siguiente
    

        print('Nuevo Cliente',siguiente)

        NombreCliente = ""
        while f_ValidaCadenaVacia(NombreCliente):
          NombreCliente=input('¿Nombre cliente?: ')
          if f_ValidaCadenaVacia(NombreCliente):
            print("Nose puede dejar vacio, ingresa un nombre")
 
        cliente[siguiente] = NombreCliente

        print("Cliente registrado correctamente") 
        print("") #Espacio

    elif (opcion=='3'):
        print(' **** Salas ****')
        print("") #Espacio
        
        siguiente = contador.get('Salas')

        # update value
        siguiente =  siguiente + 1
        contador['Salas'] = siguiente
        print('Nueva Sala',siguiente)
        

        NombreSala = ""
        while f_ValidaCadenaVacia(NombreSala):
           NombreSala=input('¿Nombre Sala?: ')
           if f_ValidaCadenaVacia(NombreSala):
             print("No se puede dejar vacio, ingresa un nombre a la sala")
             NombreSala=input('¿Nombre sala?: ')

        CapacidadSala=0


        #####Validacion de Capacidad
        while True:
          try:
             CapacidadSala=int(input(f'¿Capacidad Sala?: '))
             if  (CapacidadSala <= 0):
               print(f"Debes escribir un número mayor a 0")
               #CapacidadSala=int(input('¿Capacidad Sala?: '))  
             else:
               break
          except ValueError:
            print("Favor de ingresar un dato valido")
            continue


        #####Validacion de Capacidad  

        sala[siguiente] = NombreSala
        salaCapacidad[siguiente] = CapacidadSala

        print("Sala creada correctamente")
        print("")#Espacio

    elif (opcion=='3'):
        print(' **** Programacion de eventos x sala ****')

        #Validacion sala
        while True:
          try:
            iSala = int(input("SALA ?: "))
          except ValueError:
            print("") #Espacio
            print("Debes escribir un número.")
            continue

          if (iSala < 0):
            print("Debes escribir un número positivo.")
            continue
          else:
            break
            if (iSala in sala):
              validar = sala.get(iSala)
              print(validar)
              break
            else:
              print ("No Existe sala " , iSala)
              continue
        #Validacion sala

        #Validacion CLIENTE
        while True:
          try:
            iCliente = int(input("CLIENTE ?: "))
          except ValueError:
            print("Debes escribir un número.")
            continue

          if (iCliente < 0):
            print("Debes escribir un número positivo.")
            continue
          else:

            if (iCliente in cliente):
              validar = cliente.get(iCliente)
              print(validar)
              print("") #Espacio
              break
            else:
              print ("No Existe cliente " , iCliente)
              continue
        #Validacion CLIENTE


        #Validacion TURNO
        while True:
            iTurno = input("TURNO [M]añana [T]arde  [N]oche ?: ")

            iTurno = iTurno.upper()

            if (iTurno != "M" and iTurno != "T" and iTurno != "N"):
              print("Turno no valido debe ser [M] [T] [N]")
              continue
            else:
              break
        #Validacion TURNO

        #Validacion FECHA
        while True:
            from datetime import datetime,date,timedelta

            test_str = input("FECHA AGENDA DD-MM-YY ?: ")
 
            # initializing format
            format = "%d-%m-%y"
            EndDate = datetime.now()+timedelta(days=2)

            try:
                res = datetime.strptime(test_str, format)
                v_abuscar = test_str + str(iSala) + iTurno 

                if (res >= EndDate) and not f_EncontroCita(v_abuscar,agenda):
                  break
                else:
                      if not (res >= EndDate):
                       print("Fecha debe ser con 2 dias de Anticipacion")
                       continue
                      else:
                        print("La cita ya existe!!!")
                        continue
            except ValueError:
              #res = False
              print("Fecha Invalida ",test_str )
              continue

        #Validacion FECHA
        #8 Car fecha + 1 Car Sala +  1 Car Turno + 1 Car cliente
        AgendaFolio = test_str + str(iSala) + iTurno + str(iCliente) 
        print("Folio de agenda = ",AgendaFolio)

        AgendaEventoNombre = input("Nombre Evento ?: ")
        agenda[AgendaFolio] = AgendaEventoNombre
        print("")#Espacio

    elif (opcion=='4'):
        print(' ** Programacion de eventos x sala **')
        #print(contador)

        #Validacion sala
        while True:
          try:
            iSala = int(input("SALA ?: "))
          except ValueError:
            print("") #Espacio
            print("Debes escribir un número.")
            continue

          if (iSala < 0):
            print("Debes escribir un número positivo.")
            continue
          else:
            break
            if (iSala in sala):
              validar = sala.get(iSala)
              print(validar)
              break
            else:
              print ("No Existe sala " , iSala)
              continue
        #Validacion sala

        #Validacion CLIENTE
        while True:
          try:
            iCliente = int(input("CLIENTE ?: "))
          except ValueError:
            print("Debes escribir un número.")
            continue

          if (iCliente < 0):
            print("Debes escribir un número positivo.")
            continue
          else:

            if (iCliente in cliente):
              validar = cliente.get(iCliente)
              print(validar)
              print("") #Espacio
              break
            else:
              print ("No Existe cliente " , iCliente)
              continue
        #Validacion CLIENTE

        #Validacion TURNO
        while True:
            iTurno = input("TURNO [M]añana [T]arde  [N]oche ?: ")

            iTurno = iTurno.upper()

            if (iTurno != "M" and iTurno != "T" and iTurno != "N"):
              print("Turno no valido debe ser [M] [T] [N]")
              continue
            else:
              break
        #Validacion TURNO

        #Validacion FECHA
        while True:
            from datetime import datetime,date,timedelta
            try:
                test_str = input("FECHA AGENDA DD-MM-YY ?: ")

                format = "%d-%m-%y"
                EndDate = datetime.now()+timedelta(days=2)

                res = datetime.strptime(test_str, format)
                v_abuscar = test_str + str(iSala) + iTurno 

                if (res >= EndDate) and not f_EncontroCita(v_abuscar,agenda):
                  break
                else:
                    if not (res >= EndDate):
                      print("Fecha debe ser con 2 dias de Anticipacion")
                      continue
                    else:
                        print("La cita ya existe!!!")
                        continue
            except ValueError:
              #res = False
              print("Fecha Invalida ",test_str )
              continue

        #Validacion FECHA
        #8 Car fecha + 1 Car Sala +  1 Car Turno + 1 Car cliente
        AgendaFolio = test_str + str(iSala) + iTurno + str(iCliente) 
        print("Folio de agenda = ",AgendaFolio)

        AgendaEventoNombre = input("Nombre Evento ?: ")
        agenda[AgendaFolio] = AgendaEventoNombre

        print(agenda)
                
    elif (opcion=='9'):
        print(' ** Saliendo del menu  **')
    #else:
        #print('No existe la opcion...')